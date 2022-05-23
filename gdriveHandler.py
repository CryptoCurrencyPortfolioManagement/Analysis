import os
import io
from apiclient import discovery
from httplib2 import Http
import oauth2client
from oauth2client import file, client, tools
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload, MediaDownloadProgress
from googleapiclient.discovery import build
from tqdm import tqdm
import datetime

import pandas as pd
import openpyxl

# retrieve toekn from google to authenticate at drive api
def get_creds():
    obj = lambda: None

    lmao = {"auth_host_name":'localhost', 'noauth_local_webserver':'store_true', 'auth_host_port':[8080, 8090], 'logging_level':'ERROR'}
    for k, v in lmao.items():
        setattr(obj, k, v)

    SCOPES = 'https://www.googleapis.com/auth/drive'
    store = file.Storage('token.json')
    creds = store.get()
    if not creds or creds.invalid:
        flow = client.flow_from_clientsecrets('credentials.json', SCOPES)
        creds = tools.run_flow(flow, store, obj)

    return creds

# search files within google drive and return a dataframe of the results, implemented in two ways 1) live: retrieve information at runtime from google drive 2) cached: retrieve information from a cached file
def searchFiles(query, drive_folder_id = "1gsciAbuYI9BgwyE4b95v30qgSBJgxh5_", live= True):
    creds = get_creds()
    drive_service = build('drive', 'v3', credentials=creds)

    results = {}

    if live:
        page_token = None
        while True:
            response = drive_service.files().list(q="mimeType='application/zip' and name contains '{}' and '{}' in parents".format(query, drive_folder_id),
                                                  spaces= 'drive',
                                                  fields='nextPageToken, files(id, name)',
                                                  pageToken=page_token).execute()
            for file in response.get('files', []):
                request = drive_service.files().get_media(fileId=file.get('id'))
                fh = io.BytesIO()
                downloader = MediaIoBaseDownload(fh, request)
                done = False
                while done is False:
                    status, done = downloader.next_chunk()
                results_df = pd.DataFrame(openpyxl.load_workbook(fh)[[x.title for x in openpyxl.load_workbook(fh).worksheets][0]].values)
                results_df.columns = results_df.iloc[0].values
                results_df = results_df.iloc[1:]
                results[file.get('name')] = results_df
            page_token = response.get('nextPageToken', None)
            if page_token is None:
                break
    else:
        files = pd.read_csv("all_runs.csv", index_col=0)
        files = files.loc[files["name"].str.contains(query)].drop_duplicates(subset= ["name"])
        for i, row in files.iterrows():
            request = drive_service.files().get_media(fileId=row['id'])
            fh = io.BytesIO()
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while done is False:
                status, done = downloader.next_chunk()
            results_df = pd.DataFrame(openpyxl.load_workbook(fh)[[x.title for x in openpyxl.load_workbook(fh).worksheets][0]].values)
            results_df.columns = results_df.iloc[0].values
            results_df = results_df.iloc[1:]
            results[row["name"]] = results_df
    return results

# create file for caching the data from google drive
def getAllFiles(drive_folder_id = "1gsciAbuYI9BgwyE4b95v30qgSBJgxh5_"):
    creds = get_creds()
    drive_service = build('drive', 'v3', credentials=creds)

    results = pd.DataFrame(columns=["id", "name"])

    page_token = None
    while True:
        response = drive_service.files().list(
            q="mimeType='application/zip' and '{}' in parents".format(drive_folder_id),
            spaces='drive',
            fields='nextPageToken, files(id, name)',
            pageToken=page_token).execute()
        for file in response.get('files', []):
            results = results.append({"id": file.get('id'), "name": file.get('name')}, ignore_index=True)
        page_token = response.get('nextPageToken', None)
        results.to_csv("all_runs.csv")
        print(results.shape[0])
        if page_token is None:
            break

if __name__ == '__main__':
    getAllFiles()